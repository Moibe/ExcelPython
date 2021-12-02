# -*- coding: utf-8 -*-

"""
Created on Fri Mar 12 03:47:54 2021

@author: Moibe
#Éste programa revisa los registros de excel para crear un json con los resultados del comercio.
"""

import openpyxl
import json 
import requests
import time
import calendar
#import datetime
from datetime import date
from datetime import datetime
tiempo = calendar.timegm(time.gmtime()) #tiempoUnix.

texttiempo = str(tiempo)

print("Descarga de Json con los valores actuales. ") 

with open('rango.json') as fp:
    js_param = json.load(fp)
    
#Obtener los valores que teníamos guardados de la última revisión. 
for row in js_param: 
    
    #Valores generales:
    #Si el archivo se automatiza, la última revisión cada vez sería a las 00:00. 
    all_last_time = row['all_last_time'] #La última vez que se corrió el programa y que se actualizaron estos valores.
    adstotals_last_cell = row['adstotals_last_cell'] #La última celda revisada de Adwords Totales.
    adscampaigns_last_cell = row['adscampaigns_last_cell'] #La última celda revisada de Adwords por campaña.
    ads_spent = row['ads_spent']
    paypal_daily_last_cell = row['paypal_daily_last_cell'] #La última venta del día anterior, obtenida con el reseteador a las 00.00.
    
    # #Valores de las campañas: Cuando llevaba cada campaña desde la última revisión. 
    # valor_aleman = row['gp_aleman']
    # valor_portugues =  row['gp_portugues']
    # valor_reinoUnido = row['actual_reinoUnido']
    # valor_english = row['gp_english']
    # valor_europe = row['gp_europe']
    # valor_espanol =  row['gp_espanol']
    # valor_frances =  row['actual_frances']
    # valor_englishWorld = row['gp_englishWorld']
    
    #Valores de los productos:
    valor_en = row['paypal_en']
    valor_es = row['paypal_es']
    valor_pt = row['paypal_pt']
    valor_uk = row['paypal_uk']
    valor_fr = row['paypal_fr']
    valor_de = row['paypal_de']
    valor_eu = row['paypal_eu']
    #Valores de las cuentas: 
    valor_alpha = row['paypal_alpha']
    valor_fuel = row['paypal_fuel']
    valor_gold = row['paypal_gold']
    valor_grande = row['paypal_grande']
    valor_gsm = row['paypal_gsm']
    valor_joy = row['paypal_joy']
    valor_linked = row['paypal_linked']
    valor_telsat = row['paypal_telsat'] 
    valor_mmc = row['paypal_mmc'] 
    valor_skk = row['paypal_skk'] 
    valor_dreamnet = row['paypal_dreamnet'] 
       
    paypal_total = row['paypal_total']
    all_win = row['all_win']

time.sleep(2)

hoy = date.today() #2021-02-28
hoy_fecha_tiempo = datetime.today() #2021-02-28 23:16:42.756855
timeTag = str(hoy_fecha_tiempo)

#ADWORDS

print("Descarga y revisión de datos de Ads: ")
print("...descargando archivo desde Google. ")
#Baja una actualización del archivo de Adwords. 
#Ads
#url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vR8WxW_O5cE4IYnxH3zxQvL_4OsWR7yTOF9ixDPGHkFKsxOzmO4dIPkYE_C-oDki4rJDwMJ-aFLGkOD/pub?output=xlsx'
#AdsTest
#url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSCgnrjgdwlXI_5FxEvNxhzfDiMKFgFH2ryQ4X4xxAg1Zh-OSD4w10dfAMqDoy7grRr_UhEIxGo9zx4/pub?output=xlsx'
#Ads2022
url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vQ_GcwJ19EBB7nbat0b9QJgTflqh3YjzKIplfNcxFMoQP-y3t1Ttq5LqtZdjBBeAs6S4UPwTK5FvX0y/pub?output=xlsx'

r = requests.get(url, allow_redirects=True)

open('AdsStream.xlsx', 'wb').write(r.content)

print("Creando objetos de stream de Ads.")
#Crea los objetos con los que trabajarás el AdsStream.
wb_adsstream = openpyxl.load_workbook("AdsStream.xlsx")
#Éste Stream es para la hoja llamada Ads Totals que contiene totales generales.
ws_adstotals = wb_adsstream["Ads Totals"]
#Éste otro Stream es para la hoja default que contiene las campañas. 
ws_adscampaigns = wb_adsstream["Campaigns"]


#ADS TOTALS
#Primero obtenemos el valor general. 
#La columna B es la que contiene el costo.
columna_costo = 'B'
#Obten la nueva última fila de Ads Totals.
new_adstotals_last_cell = ws_adstotals.max_row
str_new_adstotals_last_cell = str(new_adstotals_last_cell)
celda_costo = columna_costo + str_new_adstotals_last_cell
costo_ads = ws_adstotals[celda_costo].value
print("Ésta es el costo de Ads al día de hoy.")
print(costo_ads)
time.sleep(3)

#CAMPAIGNS
#RECORRE LAS ROWS EXISTENTES ANOTANDO LAS CAMPAÑAS EN CUESTION:
#Define la celda inicial y la celda final: 
#Obtenida del Json.
celda_adscampaigns_actual =  adscampaigns_last_cell
print("Ésta es la celda inicial del recorrido de las CAMPAÑAS:")
print(celda_adscampaigns_actual)
#Obten la nueva última fila de Ads Campaigns.
new_adscampaigns_last_cell = ws_adscampaigns.max_row
celda_adscampaigns_final = new_adscampaigns_last_cell
print("Ésta es la celda FINAL del recorrido de las CAMPAÑAS:")
print(celda_adscampaigns_final)
#New rows se utiliza para saber si el reporte de Ads se ha actualizado más de una vez y si por lo tanto estaríamos sumando de más.
new_rows = celda_adscampaigns_final - celda_adscampaigns_actual

print("Éstas son las nuevas filas (new rows): ")
print(new_rows)

#La columna B contiene el nombre de la campaña y la columna C su costo.
columna_campaign = 'B'
columna_costo_campaign = 'C'

temp_aleman = 0
temp_english = 0
temp_englishWorld = 0
temp_espanol = 0
temp_europe = 0
temp_frances = 0
temp_portugues = 0
temp_reinoUnido = 0
    
for celda_adscampaigns_actual in range(celda_adscampaigns_actual, celda_adscampaigns_final + 1):
    print("Ésta es la celda_actual, el inicio del for: ")
    print(celda_adscampaigns_actual)
    
    try: 
        str_new_adscampaigns_actual = str(celda_adscampaigns_actual)
        celda_campana = columna_campaign + str_new_adscampaigns_actual
        celda_costo_campana = columna_costo_campaign + str_new_adscampaigns_actual
        campana = ws_adscampaigns[celda_campana].value
        costo_campana = ws_adscampaigns[celda_costo_campana].value
        print("Ésta es la Campaña en Cuestion: ")
        print(campana)
        print("Y éste es el costo que tuvo: ")
        print(costo_campana)
        celda_adscampaigns_actual = celda_adscampaigns_actual + 1
        time.sleep(4)
        
         #Ahora vas a sumar ese mismo valor a la campaña correspondiente:
        if "GP - English" in campana:
            print("La campaña es GP - English")
            temp_english = costo_campana
            print("Valor temporal de GP - English: ")
            print(temp_english)
        elif "GP - Español" in campana:
            print("La campaña es GP - Español")
            temp_espanol = costo_campana
            print("Valor temporal de GP - Español: ")
            print(temp_espanol)
        elif "GP - Portugués" in campana:
            print("La campaña es GP - Portugués")
            temp_portugues = costo_campana
            print("Valor temporal de GP - Portugués: ")
            print(temp_portugues)
        elif "Actual - Reino Unido" in campana:
            print("La campaña es Actual - Reino Unido")
            temp_reinoUnido = costo_campana
            print("Valor temporal de Actual - Reino Unido: ")
            print(temp_reinoUnido)
        elif "Actual - Francés" in campana:
            print("La campaña es Actual - Francés")
            temp_frances = costo_campana
            print("Valor temporal de Actual - Francés: ")
            print(temp_frances)
        elif "GP - Aleman" in campana:
            print("La campaña es GP - Alemán")
            temp_aleman = costo_campana
            print("Valor temporal de GP - Alemán: ")
            print(temp_aleman)
        elif "GP - Europe" in campana:
            print("La campaña es GP - Europe")
            temp_europe = costo_campana
            print("Valor temporal de GP - Europe: ")
            print(temp_europe)
        elif "GP - English World" in campana:
            print("La campaña es GP - English World")
            temp_englishWorld = costo_campana
            print("Valor temporal de English World: ")
            print(temp_englishWorld)
    except: 
            print("Cayó en error 1....")

time.sleep(3)

print("Ahora que ya acabo el for, puede guardar los últimos valores temporales que quedaron, como los oficiales.")
valor_english = temp_english
valor_espanol = temp_espanol
valor_portugues = temp_portugues
valor_reinoUnido = temp_reinoUnido
valor_frances = temp_frances
valor_aleman = temp_aleman
valor_europe = temp_europe
valor_englishWorld = temp_englishWorld

print("Terminamos de trabajar con los datos de ADWORDS.")
time.sleep(10)

#PAYPAL

print("Descarga y revisión de datos de Paypal...")
#Baja una actualización del archivo de ventas. 
url = 'https://www.coding-depot.dev/paypalinfosample.xlsx'
r = requests.get(url, allow_redirects=True)

open('MoneyStream.xlsx', 'wb').write(r.content)

print("Creando objetos de stream de Paypal.")
#Crea los objetos con los que trabajarás el MoneyStream.
wb_moneystream = openpyxl.load_workbook("MoneyStream.xlsx")
ws_moneystream = wb_moneystream.active

new_paypal_daily_last_cell = ws_moneystream.max_row

ultima_venta_ayer = paypal_daily_last_cell + 1 
print("Ésta es la suma de uno al paypal last cell que viene del json...")
print(ultima_venta_ayer)
time.sleep(10)

avance = new_paypal_daily_last_cell - ultima_venta_ayer
print("Nuevas ventas: ")
print(avance)

time.sleep(4)

print("Ahora vamos a repasar todas las filas nuevas del excel de paypal para sumar.")

celda_actual = paypal_daily_last_cell


columna_ganancia = 'T'
columna_producto = 'Q'
columna_cuenta = 'R'
#Aquí se anotará lo que va contabilizado hasta ahora para sumar lo demás. 
suma = paypal_total

print("El for se hará desde celda_actual a avance: ")
print("Celda Actual: ")
print(celda_actual)
print("Avance:")
print(avance)

celda_final = celda_actual + avance
print("Ésta es la celda final: ")
print(celda_final)
time.sleep(9)
print("Inicia el FOR: ")
#Se repasan todas las celdas nuevas de Moneystream. 
for celda_actual in range(celda_actual, celda_final):
    print("Ésta es la celda_actual, el inicio del for: ")
    print(celda_actual)
    
    try: 
        #Definición del trío de celdas con las que trabajaremos. 
        fila_str = str(celda_actual)
        print("Fila: " + fila_str)
        celda_money = columna_ganancia + fila_str
        celda_producto = columna_producto + fila_str
        celda_cuenta = columna_cuenta + fila_str
        print("Ésta es la celda del $: ")
        print(celda_money)
        print("Ésta es la celda del producto: ")
        print(celda_producto)
        print("Ésta es la celda de la cuenta: ")
        print(celda_cuenta)
        print("Éste es el dinero obtenido: ")
        valor = ws_moneystream[celda_money].value
        print(valor)
        print("Éste es el producto contabilizado:")
        producto = ws_moneystream[celda_producto].value
        print(producto)
        print("Ésta es la cuenta que lo recibió:")
        cuenta = ws_moneystream[celda_cuenta].value
        print(cuenta)
        
        #GANANCIAS DE PAYPAL POR CAMPAÑA
        #Ahora vas a sumar ese mismo valor al producto correspondiente:
        if "EN" in producto:
            print("El producto es ENGLISH!")
            valor_en = valor_en + valor
            print("Valor EN so far: ")
            print(valor_en)
        elif "ES" in producto:
                print("El producto es ESPAÑOL!")
                valor_es = valor_es + valor
                print("Valor ES so far: ")
                print(valor_es)
        elif "PT" in producto:
                    print("El producto es Portugués!")
                    valor_pt = valor_pt + valor
                    print("Valor PT so far: ")
                    print(valor_pt)
        elif "UK" in producto:
                        print("El producto es Reino Unido")
                        valor_uk = valor_uk + valor
                        print("Valor UK so far: ")
                        print(valor_uk)
        elif "FR" in producto:
                            print("El producto es Francés!")
                            valor_fr = valor_fr + valor
                            print("Valor FR so far: ")
                            print(valor_fr)
        elif "DE" in producto:
                                print("El producto es Alemán!")
                                valor_de = valor_de + valor
                                print("Valor DE so far: ")
                                print(valor_de)
        elif "EU" in producto:
                                    print("El producto es Europe!")
                                    valor_eu = valor_eu + valor
                                    print("Valor EU so far: ")
                                    print(valor_eu)
        
        
        #GANANCIAS DE PAYPAL POR CUENTA.                            
        #Y vas ahora a sumar ese mismo valor a la cuenta correspondiente: 
        #Ahora vas a sumar ese mismo valor al producto correspondiente:
        if "Alpha" in cuenta:
            print("La cuenta es Alpha Summit: ")
            valor_alpha = valor_alpha + valor
            print("Valor Alpha so far: ")
            print(valor_alpha)
        elif "fuelcom" in cuenta:
            print("La cuenta es Fuelcom: ")
            valor_fuel = valor_fuel + valor
            print("Valor Fuelbank so far: ")
            print(valor_fuel)
        elif "Gold" in cuenta:
            print("La cuenta es Moibe Gold: ")
            valor_gold = valor_gold + valor
            print("Valor MoibeGold so far: ")
            print(valor_gold)
        elif "GrandeGold" in cuenta:
            print("La cuenta es Grande Gold: ")
            valor_grande = valor_grande + valor
            print("Valor Grande Gold so far: ")
            print(valor_grande)
        elif "GSM" in cuenta:
            print("La cuenta es GSM : ")
            valor_gsm = valor_gsm + valor
            print("Valor GSM so far: ")
            print(valor_gsm)
        elif "JoyToei" in cuenta:
            print("La cuenta es JoyToei: ")
            valor_joy = valor_joy + valor
            print("Valor JoyToei so far: ")
            print(valor_joy)
        elif "linkedsys" in cuenta:
            print("La cuenta es LinkedSys: ")
            valor_linked = valor_linked + valor
            print("Valor LinkedSys so far: ")
            print(valor_linked)
        elif "telsat" in cuenta:
            print("La cuenta es Telsat: ")
            valor_telsat = valor_telsat + valor
            print("Valor Telsat so far: ")
            print(valor_telsat)
        elif "MMC" in cuenta:
            print("La cuenta es MMC: ")
            valor_telsat = valor_telsat + valor
            print("Valor MMC so far: ")
            print(valor_mmc)
        elif "SKK" in cuenta:
            print("La cuenta es Skk: ")
            valor_telsat = valor_telsat + valor
            print("Valor Skk so far: ")
            print(valor_skk)
        elif "Dreamnet" in cuenta:
            print("La cuenta es Dreamnet: ")
            valor_dreamnet = valor_dreamnet + valor
            print("Valor Dreamnet so far: ")
            print(valor_dreamnet)
       
        suma = suma + valor
        print("Valor total so FAR: ")
        print(suma)
        celda_actual = celda_actual + 1
    except: 
            print("Cayó en error 2....")
      
#TOTALES
#Totales Generales:
print("La suma total da:")
print(suma)
#Totales por Producto:
print("La suma de EN:")
print(valor_en)
print("La suma de ES:")
print(valor_es)
print("La suma de PT:")
print(valor_pt)
print("La suma de UK:")
print(valor_uk)
print("La suma de FR:")
print(valor_fr)
print("La suma de DE:")
print(valor_de)
print("La suma de EU:")
print(valor_eu)
#Totales por Cuenta: 
print("La suma de Alpha Summit:")
print(valor_alpha)
print("La suma de Fuelbank:")
print(valor_fuel)
print("La suma de Moibe Gold:")
print(valor_gold)
print("La suma de Grande Gold:")
print(valor_grande)
print("La suma de GSM:")
print(valor_gsm)
print("La suma de JoyToei:")
print(valor_joy)
print("La suma de LinkedSys:")
print(valor_linked)
print("La suma de Telsat:")
print(valor_telsat)


#Ahora vamos a comparar lo que se ha gastado contra lo que se ha recibido. 
ganancia = suma - costo_ads
print("Hemos ganado:")
print(ganancia)
time.sleep(15)

#ACTUALIZACIÓN DE JSON
for row in js_param:
    
    print("Estos serán los nuevos valores a guardar: ")
    row['all_last_time'] = timeTag
    row['adstotals_last_cell'] = new_adstotals_last_cell
    row['adscampaigns_last_cell'] = new_adscampaigns_last_cell
    #Ya no necesito acualizar la última celda porque cada vez repasará desde la primera de paypal del día.
    #Ahora se actualizará vía el reseteador.
    #row['paypal_daily_last_cell'] = new_paypal_daily_last_cell
    row['ads_spent'] = costo_ads
    row['paypal_total'] = suma
    #Valores para las campañas
    row['gp_english'] = valor_english
    row['gp_espanol'] = valor_espanol
    row['gp_aleman'] = valor_aleman
    row['actual_reinoUnido'] = valor_reinoUnido
    row['gp_englishWorld'] = valor_englishWorld
    row['actual_frances'] = valor_frances
    row['gp_portugues'] = valor_portugues
    row['gp_europe'] = valor_europe
    #Valores por producto:
    row['paypal_en'] = valor_en
    row['paypal_es'] = valor_es
    row['paypal_pt'] = valor_pt
    row['paypal_uk'] = valor_uk
    row['paypal_fr'] = valor_fr
    row['paypal_de'] = valor_de
    row['paypal_eu'] = valor_eu
    #Valores por cuenta: 
    row['paypal_alpha'] = valor_alpha
    row['paypal_fuel'] = valor_fuel
    row['paypal_gold'] = valor_gold
    row['paypal_grande'] = valor_grande
    row['paypal_gsm'] = valor_gsm
    row['paypal_joy'] = valor_joy
    row['paypal_linked'] = valor_linked
    row['paypal_telsat'] = valor_telsat

    print("Y así queda nuestro nuevo set:")
    print(js_param)
        
    #Escribe 
    with open('rango.json', 'w', encoding='utf-8') as jsonf: 
        jsonf.write(json.dumps(js_param, indent=0)) 